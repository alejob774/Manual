# -*- coding: utf-8 -*-
"""
Enriquecedor de Concesionarios (Motor + GUI)
--------------------------------------------
• Hoja 1 → Tabla con CÓDIGO (10 dígitos), PEDIDO, NOTA.
• Hoja 2 → Base donde se insertarán PERIODO, PEDIDO, NOTA.
• PERIODO se ingresa manualmente en la UI (Opción C — sin validación).
• Se eliminan filas que no tengan ambos (PEDIDO y NOTA).
• Resultado se guarda en /Resultado/<nombre_original>.xlsx
"""

import sys
import re
import os
import unicodedata
import math
from typing import Optional, Tuple, Dict, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

# ==========================================================================
# UTILIDADES GENERALES
# ==========================================================================

CODE_RE = re.compile(r"^\d{10}$")

def strip_accents_lower(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.lower().strip()

def only_digits(s) -> Optional[str]:
    if s is None:
        return None
    digits = re.sub(r"\D", "", str(s))
    return digits if len(digits) == 10 else None

def is_10digit_code(val) -> bool:
    code = only_digits(val)
    return code is not None and CODE_RE.match(code) is not None

def last_used_col(ws: Worksheet) -> int:
    max_c = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in ("", None):
                max_c = max(max_c, cell.column)
    return max_c

def last_used_row(ws: Worksheet) -> int:
    max_r = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in ("", None):
                max_r = max(max_r, cell.row)
    return max_r

def row_nonempty_cols(ws: Worksheet, row_idx: int) -> List[int]:
    return [c for c in range(1, ws.max_column + 1)
            if ws.cell(row=row_idx, column=c).value not in ("", None)]

def row_has_any_code(ws: Worksheet, row_idx: int) -> bool:
    return any(is_10digit_code(ws.cell(row=row_idx, column=c).value)
               for c in range(1, ws.max_column + 1))

def detect_header_heuristic(ws: Worksheet) -> int:
    """
    Heurística robusta para detectar la fila de encabezado en Hoja 2.

    - Busca la primera fila con algún código de 10 dígitos (datos).
    - Entre la fila 1 y la fila anterior a esa, elige la que:
        * tiene más celdas no vacías
        * tiene más valores distintos (evita filas como 'Extracto GMF' repetido)
        * tiene mayor proporción de texto (letras)
    Si no encuentra códigos, busca en las primeras ~30 filas con datos.
    """
    max_r = ws.max_row
    max_c = ws.max_column

    first_code_row = None
    last_data_row = 0
    metrics = {}  # fila -> (nonempty, distinct, text_ratio)

    for r in range(1, max_r + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
        nonempty = [v for v in row_vals if v not in ("", None)]
        if nonempty:
            last_data_row = r

        # ¿la fila tiene algún código de 10 dígitos?
        if any(is_10digit_code(v) for v in nonempty) and first_code_row is None:
            first_code_row = r

        if nonempty:
            # métrica de texto / variedad
            text_cnt = 0
            for v in nonempty:
                s = str(v)
                if re.search(r"[A-Za-zÁÉÍÓÚÑáéíóúñ]", s):
                    text_cnt += 1
            text_ratio = text_cnt / len(nonempty)
            distinct = len(set(nonempty))
            metrics[r] = (len(nonempty), distinct, text_ratio)

    if first_code_row is None or first_code_row <= 1:
        search_upto = min(30, last_data_row)
    else:
        search_upto = first_code_row - 1

    best_row = 1
    best_score = (-1, -1, -1.0)  # (nonempty, distinct, text_ratio)

    for r in range(1, search_upto + 1):
        if r not in metrics:
            continue
        score = metrics[r]
        if score > best_score:
            best_score = score
            best_row = r

    print(f"[Diag] Hoja2: encabezado heurístico = fila {best_row}, score={best_score}, first_code_row={first_code_row}, last_data_row={last_data_row}")
    return best_row

# ==========================================================================
# HOJA 1 – MAPEO DE CÓDIGOS
# ==========================================================================

def find_code_and_name_block(ws: Worksheet):
    print("[Diag] Hoja1: buscando bloque Código + Nombre...")
    col_candidates = {}

    for c in range(1, ws.max_column + 1):
        rows_code = [r for r in range(1, ws.max_row + 1)
                     if is_10digit_code(ws.cell(row=r, column=c).value)]
        if rows_code:
            col_candidates[c] = rows_code

    if not col_candidates:
        print("[Diag] Hoja1: no hay códigos.")
        return None

    def contiguous(lst):
        lst = sorted(lst)
        out = []
        start = lst[0]
        prev = lst[0]
        for x in lst[1:]:
            if x == prev + 1:
                prev = x
            else:
                out.append((start, prev))
                start = x
                prev = x
        out.append((start, prev))
        return out

    best = None
    for c_code, rows in col_candidates.items():
        for r0, r1 in contiguous(rows):
            length = r1 - r0 + 1
            best_name = None
            best_q = -1
            for c_name in [c_code - 1, c_code + 1]:
                if 1 <= c_name <= ws.max_column:
                    filled = 0
                    text_cnt = 0
                    for r in range(r0, r1 + 1):
                        v = ws.cell(row=r, column=c_name).value
                        if v not in ("", None):
                            filled += 1
                            if not is_10digit_code(v) and re.search(r"[A-Za-z]", str(v)):
                                text_cnt += 1
                    if filled > 0:
                        q = text_cnt / filled
                        if q > best_q:
                            best_q = q
                            best_name = c_name
            if best_name:
                score = (length, best_q)
                if (best is None) or score > (best[0], best[5]):
                    best = (length, r0, r1, c_code, best_name, best_q)

    if not best:
        return None

    _, r0, r1, c_code, c_name, _ = best
    print(f"[Diag] Hoja1: bloque {r0}-{r1} → Código:{c_code}, Nombre:{c_name}")
    return r0, r1, c_code, c_name

def find_pedido_nota_columns(ws: Worksheet, r0, r1, base_col):
    hdr = r0 - 1
    if hdr < 1:
        return None

    def H(c):
        return strip_accents_lower(ws.cell(row=hdr, column=c).value)

    for c in range(base_col + 1, ws.max_column):
        if ("pedido" in H(c) and "nota" in H(c+1)) and ("clase" not in H(c)):
            print(f"[Diag] Hoja1: Pedido/Nota = {c}/{c+1}")
            return c, c+1
    return None

def find_pedido_nota_header_anywhere(ws: Worksheet):
    """
    Busca en toda la hoja una fila donde haya columnas consecutivas
    PEDIDO(S) y NOTA. Devuelve (fila_header, col_pedido, col_nota)
    o None si no encuentra nada.
    """
    print("[Diag] Hoja1: buscando encabezados globales PEDIDO/NOTA...")
    max_c = ws.max_column
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_c):
            h1 = strip_accents_lower(ws.cell(row=r, column=c).value)
            h2 = strip_accents_lower(ws.cell(row=r, column=c+1).value)
            if ("pedido" in h1 and "nota" in h2) and ("clase" not in h1):
                print(f"[Diag] Hoja1: Encabezado PEDIDO/NOTA encontrado en fila {r}, columnas {c}/{c+1}")
                return r, c, c+1
    print("[Diag] Hoja1: no se encontró encabezado PEDIDO/NOTA global.")
    return None

def find_code_col_near_header(ws: Worksheet, hdr_row: int):
    """
    Dado el número de fila del encabezado (hdr_row), busca la mejor columna
    de CÓDIGO justo debajo del encabezado, midiendo la longitud del bloque
    contiguo de filas con códigos de 10 dígitos a partir de hdr_row+1.
    Ignora las columnas cuyo encabezado sea PEDIDO/NOTA.
    """
    best_c = None
    best_run = 0

    for c in range(1, ws.max_column + 1):
        h = strip_accents_lower(ws.cell(row=hdr_row, column=c).value)
        # Nunca usar las columnas de PEDIDO/NOTA como columna de código
        if "pedido" in h or "nota" in h:
            continue

        run = 0
        r = hdr_row + 1
        while r <= ws.max_row and is_10digit_code(ws.cell(row=r, column=c).value):
            run += 1
            r += 1

        if run > best_run:
            best_run = run
            best_c = c

    print(f"[Diag] Hoja1: columna de CÓDIGO (modo encabezados) = {best_c}, filas contiguas con código = {best_run}")
    return best_c, best_run

def build_code_to_pedido_nota(ws: Worksheet) -> Dict[str, Tuple[Optional[str], Optional[str]]]:
    """
    Construye un mapeo CÓDIGO (10 dígitos) -> (PEDIDO, NOTA).

    MODO 1 (preferido): detecta explícitamente el encabezado PEDIDOS / Nota
    en toda la hoja (caso reportes tipo ASONAC / tablas pivote) y usa solo
    el bloque contiguo de la tabla resumen inmediatamente debajo.

    MODO 2 (fallback): si lo anterior falla, usa la lógica original basada
    en encontrar un bloque largo de CÓDIGO + NOMBRE.
    """
    # -----------------------------
    # MODO 1: encabezados globales
    # -----------------------------
    hdr_info = find_pedido_nota_header_anywhere(ws)
    if hdr_info:
        hdr_row, c_p, c_n = hdr_info
        code_col, run_len = find_code_col_near_header(ws, hdr_row)

        if code_col and run_len > 0:
            mapping: Dict[str, Tuple[Optional[str], Optional[str]]] = {}
            start = hdr_row + 1
            end   = start + run_len - 1

            for r in range(start, end + 1):
                code = only_digits(ws.cell(row=r, column=code_col).value)
                if not code:
                    continue

                pedido = ws.cell(row=r, column=c_p).value
                nota   = ws.cell(row=r, column=c_n).value

                # Solo considerar filas donde haya al menos PEDIDO o NOTA
                if pedido in ("", None) and nota in ("", None):
                    continue

                mapping[code] = (
                    str(pedido).strip() if pedido not in ("", None) else None,
                    str(nota).strip()   if nota   not in ("", None) else None
                )

            print(f"[Diag] Hoja1: mapeo creado (modo encabezados) para {len(mapping)} códigos.")
            if mapping:
                return mapping
        else:
            print("[Diag] Hoja1: no se pudo determinar columna de CÓDIGO en modo encabezados.")

    # -----------------------------
    # MODO 2: lógica original (fallback)
    # -----------------------------
    print("[Diag] Hoja1: usando lógica de bloque Código+Nombre (fallback).")
    block = find_code_and_name_block(ws)
    if not block:
        return {}

    r0, r1, c_code, c_name = block
    pn = find_pedido_nota_columns(ws, r0, r1, max(c_code, c_name))
    if not pn:
        print("[Diag] Hoja1: no hay columnas Pedido/Nota (modo bloque).")
        return {}

    c_p, c_n = pn
    mapping: Dict[str, Tuple[Optional[str], Optional[str]]] = {}

    for r in range(r0, r1 + 1):
        code = only_digits(ws.cell(row=r, column=c_code).value)
        if code:
            pedido = ws.cell(row=r, column=c_p).value
            nota   = ws.cell(row=r, column=c_n).value
            mapping[code] = (
                str(pedido).strip() if pedido not in ("", None) else None,
                str(nota).strip()   if nota   not in ("", None) else None
            )

    print(f"[Diag] Hoja1: mapeo creado (modo bloque) para {len(mapping)} códigos.")
    return mapping

# ==========================================================================
# HOJA 2 – AUTOFILTER
# ==========================================================================

def clear_autofilter(ws: Worksheet):
    try:
        if ws.auto_filter and ws.auto_filter.ref:
            print(f"[Diag] AutoFilter encontrado (ref={ws.auto_filter.ref}) → limpiando...")
            ws.auto_filter.ref = None
    except:
        pass

def apply_autofilter(ws: Worksheet):
    last_col = ws.max_column
    last_row = ws.max_row
    try:
        ref = f"A1:{get_column_letter(last_col)}{last_row}"
        ws.auto_filter.ref = ref
        print(f"[Diag] AutoFilter aplicado → {ref}")
    except:
        print("[Diag] No se pudo aplicar AutoFilter.")

# ==========================================================================
# HOJA 2 – DETECCIÓN DE TABLA
# ==========================================================================

def detect_table_by_last_row(ws: Worksheet):
    """
    Usa la heurística de encabezado:
    - En PROVISION REBATE GMF FEBRERO 2026 -> detecta fila 3.
    - En LIQUIDACION PAC Q4 2025 - 2DA -> detecta fila 1.
    Devuelve (fila_encabezado, columnas_no_vacías_en_el_encabezado).
    """
    hdr = detect_header_heuristic(ws)
    cols = [c for c in range(1, ws.max_column + 1)
            if ws.cell(row=hdr, column=c).value not in ("", None)]

    print(f"[Diag] Hoja2: encabezado detectado en fila {hdr}, columnas={cols[:12]}")
    return hdr, cols

def move_table_to_top(ws: Worksheet, hdr: int):
    if hdr > 1:
        ws.delete_rows(1, hdr - 1)

# ==========================================================================
# HOJA 2 – INSERCIÓN A/B/C + CRUCE
# ==========================================================================

def insert_first_three_columns(ws: Worksheet):
    ws.insert_cols(1, 3)
    ws.cell(row=1, column=1, value="PERIODO")
    ws.cell(row=1, column=2, value="PEDIDO")
    ws.cell(row=1, column=3, value="NOTA")

def locate_code_col(ws: Worksheet, mapping_keys: set):
    max_r = ws.max_row
    best_c = None
    best_overlap = -1

    for c in range(1, ws.max_column + 1):
        codes = []
        for r in range(2, max_r + 1):
            code = only_digits(ws.cell(row=r, column=c).value)
            if code:
                codes.append(code)
        if not codes:
            continue
        overlap = sum(1 for x in codes if x in mapping_keys)
        if overlap > best_overlap:
            best_overlap = overlap
            best_c = c

    print(f"[Diag] Columna de código seleccionada: {best_c}, overlap={best_overlap}")
    return best_c

def fill_pedido_nota(ws: Worksheet, mapping, code_col):
    max_r = ws.max_row
    cnt = 0
    for r in range(2, max_r + 1):
        code = only_digits(ws.cell(row=r, column=code_col).value)
        if code:
            pedido, nota = mapping.get(code, (None, None))
            if pedido:
                ws.cell(row=r, column=2, value=pedido)
            if nota:
                ws.cell(row=r, column=3, value=nota)
            if pedido and nota:
                cnt += 1
    print(f"[Diag] Pedido/Nota rellenados en {cnt} filas.")

# ==========================================================================
# HOJA 2 – ELIMINAR FILAS INVÁLIDAS
# ==========================================================================

def prune_rows_without_both(ws: Worksheet):
    """
    Misma lógica:
    - Elimina filas donde PEDIDO (col 2) o NOTA (col 3) estén vacíos.
    Optimizada:
    - Reescribe solo las filas válidas al inicio (sin borrar dentro del bucle).
    - Al final borra de una sola vez el bloque de filas sobrantes.
    """
    max_r = ws.max_row
    max_c = ws.max_column

    write_row = 2  # primera fila de datos
    deleted = 0

    for r in range(2, max_r + 1):
        p = ws.cell(row=r, column=2).value
        n = ws.cell(row=r, column=3).value

        empty_p = p in (None, "", " ")
        empty_n = n in (None, "", " ")

        if empty_p or empty_n:
            deleted += 1
            continue

        if write_row != r:
            for c in range(1, max_c + 1):
                ws.cell(row=write_row, column=c).value = ws.cell(row=r, column=c).value
        write_row += 1

    # Borrado en bloque de las filas restantes (todas inválidas)
    if write_row <= max_r:
        ws.delete_rows(write_row, max_r - write_row + 1)

    print(f"[Diag] Filas eliminadas: {deleted}")

def prune_rows_by_column(ws: Worksheet, header_row: int, col_idx: int):
    """
    Elimina filas donde la columna indicada tenga valores "vacíos":
    - None, "", "-", "nan" (string)
    - 0 o NaN (numérico)
    """
    last_r = last_used_row(ws)
    to_delete = []

    for r in range(header_row + 1, last_r + 1):
        v = ws.cell(row=r, column=col_idx).value
        delete = False

        if v is None:
            delete = True
        elif isinstance(v, str):
            s = v.strip().lower()
            if s in ("", "-", "nan", ""):
                delete = True
        elif isinstance(v, (int, float)):
            try:
                if float(v) == 0 or (isinstance(v, float) and math.isnan(v)):
                    delete = True
            except Exception:
                pass

        if delete:
            to_delete.append(r)

    for r in reversed(to_delete):
        ws.delete_rows(r, 1)

    print(f"[Diag] Filas eliminadas por filtro de columna índice {col_idx}: {len(to_delete)}")

# ==========================================================================
# HOJA 2 – PERIODO MANUAL
# ==========================================================================

def fill_periodo_manual(ws: Worksheet, periodo: str):
    max_r = ws.max_row
    for r in range(2, max_r + 1):
        ws.cell(row=r, column=1, value=periodo)
    print(f"[Diag] PERIODO '{periodo}' aplicado.")

# ==========================================================================
# MOTOR PRINCIPAL
# ==========================================================================

def process_file(input_path: str, periodo_manual: str, filter_column_name: Optional[str] = None) -> str:

    if not os.path.exists(input_path):
        raise FileNotFoundError("Archivo no encontrado.")

    wb = load_workbook(input_path, data_only=True)
    if len(wb.sheetnames) < 2:
        raise ValueError("Se requieren al menos 2 hojas.")

    ws1 = wb[wb.sheetnames[0]]
    ws2 = wb[wb.sheetnames[1]]

    mapping = build_code_to_pedido_nota(ws1)

    hdr, _ = detect_table_by_last_row(ws2)
    clear_autofilter(ws2)
    move_table_to_top(ws2, hdr)
    insert_first_three_columns(ws2)

    # Filtro opcional por columna (valores vacíos / 0 / "-")
    if filter_column_name:
        header_row = 1
        col_idx = None
        for c in range(1, ws2.max_column + 1):
            v = ws2.cell(row=header_row, column=c).value
            if v not in ("", None) and str(v).strip() == filter_column_name.strip():
                col_idx = c
                break

        if col_idx is not None:
            prune_rows_by_column(ws2, header_row, col_idx)
        else:
            print(f"[Diag] No se encontró la columna '{filter_column_name}' en Hoja 2; no se aplicó el filtro.")

    code_col = locate_code_col(ws2, set(mapping.keys()))
    if code_col:
        fill_pedido_nota(ws2, mapping, code_col)

    fill_periodo_manual(ws2, periodo_manual)

    prune_rows_without_both(ws2)

    apply_autofilter(ws2)

    base_dir = os.path.dirname(input_path)
    base_name = os.path.splitext(os.path.basename(input_path))[0]

    result_dir = os.path.join(base_dir, "Resultado")
    os.makedirs(result_dir, exist_ok=True)

    output_path = os.path.join(result_dir, f"{base_name}.xlsx")
    wb.save(output_path)

    print(f"[OK] Archivo generado: {output_path}")
    return output_path

# ==========================================================================
# UI (Tkinter)
# ==========================================================================

def launch_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    def select_file():
        p = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not p:
            return

        file_var.set(p)

        # Cargar encabezados de Hoja 2 y llenar el menú desplegable
        try:
            wb = load_workbook(p, data_only=True)
            if len(wb.sheetnames) < 2:
                messagebox.showwarning(
                    "Advertencia",
                    "El archivo no tiene al menos 2 hojas. "
                    "No se pudieron cargar columnas para el filtro."
                )
                return

            ws2 = wb[wb.sheetnames[1]]
            hdr, _ = detect_table_by_last_row(ws2)

            headers = []
            for c in range(1, ws2.max_column + 1):
                val = ws2.cell(row=hdr, column=c).value
                if val not in ("", None):
                    headers.append(str(val))

            menu = column_option["menu"]
            menu.delete(0, "end")

            if headers:
                for h in headers:
                    menu.add_command(
                        label=h,
                        command=lambda v=h: filter_column_var.set(v)
                    )
                # valor por defecto: primera columna encontrada
                filter_column_var.set(headers[0])
            else:
                filter_column_var.set("")

        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudieron leer las columnas de la Hoja 2:\n{e}"
            )

    def run_process():
        path = file_var.get().strip()
        per = periodo_var.get().strip()

        if path == "":
            messagebox.showerror("Error", "Selecciona un archivo Excel.")
            return
        if per == "":
            messagebox.showerror("Error", "Debes ingresar un PERIODO.")
            return

        use_filter = filter_enabled_var.get()
        selected_col = filter_column_var.get().strip()

        if use_filter and not selected_col:
            messagebox.showerror(
                "Error",
                "Selecciona una columna para aplicar el filtro o desmarca la opción."
            )
            return

        try:
            out = process_file(
                path,
                per,
                selected_col if use_filter else None
            )
            messagebox.showinfo("Éxito", f"Archivo generado:\n{out}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    root = tk.Tk()
    root.title("Enriquecedor de Concesionarios")
    root.geometry("750x520")
    root.resizable(False, False)

    tk.Label(
        root,
        text="Enriquecedor de Concesionarios",
        font=("Segoe UI", 16, "bold")
    ).pack(pady=10)

    info = (
        "REQUISITOS:\n"
        "• Sheet 1 → CÓDIGO de concesionario, PEDIDO y NOTA\n"
        "• Sheet 2 → Base a enriquecer\n"
        "• PERIODO se ingresa manualmente (Opción C – sin validación)\n"
        "• Se eliminan filas sin ambos valores PEDIDO y NOTA\n"
        "• Archivo final quedará en carpeta: Resultado/"
    )

    tk.Label(
        root,
        text=info,
        justify="left",
        wraplength=700,
        font=("Segoe UI", 10)
    ).pack(pady=10)

    file_var = tk.StringVar()
    periodo_var = tk.StringVar()

    tk.Entry(root, textvariable=file_var, width=70).pack()
    tk.Button(root, text="Buscar archivo Excel", command=select_file).pack(pady=10)

    tk.Label(
        root,
        text="PERIODO (texto libre, ejemplo 202512):",
        font=("Segoe UI", 10, "bold")
    ).pack()
    tk.Entry(root, textvariable=periodo_var, width=20).pack(pady=5)

    # ---- NUEVA SECCIÓN: filtro opcional por columna ----
    tk.Label(
        root,
        text="Filtro opcional por columna en Hoja 2\n"
             "(elimina filas con valores vacíos / 0 / \"-\")",
        font=("Segoe UI", 10, "bold")
    ).pack(pady=(15, 2))

    filter_enabled_var = tk.BooleanVar(value=False)
    tk.Checkbutton(
        root,
        text="Aplicar filtro por columna",
        variable=filter_enabled_var
    ).pack()

    tk.Label(
        root,
        text="Columna de Hoja 2 para filtrar (si el filtro está activo):",
        font=("Segoe UI", 10)
    ).pack(pady=(5, 2))

    filter_column_var = tk.StringVar()
    filter_column_var.set("")

    column_option = tk.OptionMenu(root, filter_column_var, "")
    column_option.pack(pady=5)

    # Botón principal
    tk.Button(
        root,
        text="Procesar archivo",
        font=("Segoe UI", 12),
        command=run_process
    ).pack(pady=20)

    root.mainloop()

# ==========================================================================
# ENTRYPOINT
# ==========================================================================

if __name__ == "__main__":
    if len(sys.argv) == 3:
        process_file(sys.argv[1], sys.argv[2])
    else:
        launch_gui()